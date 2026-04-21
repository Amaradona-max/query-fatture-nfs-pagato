from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
import pytest

from app.services.file_processor import NFSFTFileProcessor, PisaFTFileProcessor, PisaRicevuteFTFileProcessor, CompareFTFileProcessor


@pytest.fixture
def sample_dataframe():
    return pd.DataFrame(
        {
            "C_NOME": ["ACME Inc", "Test Corp", "ACME Inc"],
            "FAT_DATDOC": ["2025-01-01", "2025-01-02", "2025-01-01"],
            "FAT_NDOC": ["F001", "F002", "F001"],
            "FAT_DATREG": ["2025-01-01", "2025-01-02", "2025-01-01"],
            "FAT_PROT": ["EP", "P", "EP"],
            "FAT_NUM": [1, 2, 1],
            "IMPONIBILE": [100.0, 200.0, 100.0],
            "FAT_TOTFAT": [122.0, 244.0, 122.0],
            "FAT_TOTIVA": [22.0, 44.0, 22.0],
            "RA_IMPON": [100.0, 200.0, 100.0],
            "RA_CODTRIB": ["I9", "XX", "RO"],
            "RA_IMPOSTA": [5.0, 10.0, 5.0],
            "TMC_G8": ["ID1", "ID2", "ID1"],
        }
    )


def test_validate_file_success(sample_dataframe):
    processor = NFSFTFileProcessor()
    processor.validate_file(sample_dataframe)


def test_validate_file_missing_columns():
    processor = NFSFTFileProcessor()
    df = pd.DataFrame({"WRONG_COL": [1, 2]})
    with pytest.raises(ValueError, match="Colonne mancanti"):
        processor.validate_file(df)


def test_process_file_removes_duplicates(sample_dataframe, tmp_path: Path):
    processor = NFSFTFileProcessor()
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"

    sample_dataframe.to_excel(input_path, index=False)
    stats = processor.process_file(input_path, output_path)

    assert output_path.exists()
    assert stats["total_records"] == 2
    assert stats["duplicates_removed"] == 1
    assert stats["fase2_records"] == 1
    assert stats["fase3_records"] == 1


def test_process_file_pisa_splits_by_sdi(tmp_path: Path):
    df = pd.DataFrame(
        [
            {
                "Creditore": "Ragione A",
                "Numero fattura": "F001",
                "Identificativo SDI": "",
                "Data emissione": "2025-01-10",
                "Importo fattura": 120.0,
                "Importo pagato": 100.0,
            },
            {
                "Creditore": "Ragione B",
                "Numero fattura": "F002",
                "Identificativo SDI": "123",
                "Data emissione": "",
                "Importo fattura": 220.0,
                "Importo pagato": 200.0,
            },
            {
                "Creditore": "Ragione C",
                "Numero fattura": "F003",
                "Identificativo SDI": None,
                "Data emissione": "2025-02-01",
                "Importo fattura": 320.0,
                "Importo pagato": 300.0,
            },
        ],
    )

    processor = PisaFTFileProcessor()
    input_path = tmp_path / "input_pisa.xlsx"
    output_path = tmp_path / "output_pisa.xlsx"
    df.to_excel(input_path, index=False)

    stats = processor.process_file(input_path, output_path)

    assert output_path.exists()
    assert stats["total_records"] == 3
    assert stats["fase2_records"] == 2
    assert stats["fase3_records"] == 1

    wb = load_workbook(output_path, data_only=True)
    cartacee_ws = wb["Fatture Cartacee"]
    elettroniche_ws = wb["Fatture Elettroniche"]

    assert cartacee_ws["A1"].value == "NUMERO TOTALE"
    assert cartacee_ws["B1"].value == "IMPORTO"
    assert cartacee_ws["A2"].value == 2
    assert cartacee_ws["B2"].value == 400.0

    assert elettroniche_ws["A1"].value == "NUMERO TOTALE"
    assert elettroniche_ws["B1"].value == "IMPORTO"
    assert elettroniche_ws["A2"].value == 1
    assert elettroniche_ws["B2"].value == 200.0


def test_process_file_pisa_ricevute_splits_by_sdi(tmp_path: Path):
    df = pd.DataFrame(
        [
            {
                "Creditore": "Ragione A",
                "Numero fattura": "F001",
                "Data emissione": "2025-01-10",
                "Data documento": "2025-01-10",
                "Data pagamento": "",
                "IVA": "22,00",
                "Importo fattura": "122,00",
                "Identificativo SDI": "",
            },
            {
                "Creditore": "Ragione B",
                "Numero fattura": "F002",
                "Data emissione": "2025-01-11",
                "Data documento": "2025-01-11",
                "Data pagamento": "",
                "IVA": "44,00",
                "Importo fattura": "244,00",
                "Identificativo SDI": "123",
            },
            {
                "Creditore": "Ragione C",
                "Numero fattura": "F003",
                "Data emissione": "2025-01-12",
                "Data documento": "2025-01-12",
                "Data pagamento": "",
                "IVA": "0",
                "Importo fattura": "100",
                "Identificativo SDI": None,
            },
        ]
    )

    processor = PisaRicevuteFTFileProcessor()
    input_path = tmp_path / "input_pisa_ricevute.xlsx"
    output_path = tmp_path / "output_pisa_ricevute.xlsx"
    df.to_excel(input_path, index=False)

    stats = processor.process_file(input_path, output_path)

    assert output_path.exists()
    assert stats["total_records"] == 3
    assert stats["fase2_records"] == 2
    assert stats["fase3_records"] == 1


def test_process_file_pisa_ricevute_treats_zero_sdi_as_elettronica(tmp_path: Path):
    df = pd.DataFrame(
        [
            {
                "Creditore": "Ragione A",
                "Numero fattura": "F001",
                "Data emissione": "2025-01-10",
                "Data documento": "2025-01-10",
                "Data pagamento": "",
                "IVA": "22,00",
                "Importo fattura": "122,00",
                "Identificativo SDI": "0,0",
            },
            {
                "Creditore": "Ragione B",
                "Numero fattura": "F002",
                "Data emissione": "2025-01-11",
                "Data documento": "2025-01-11",
                "Data pagamento": "",
                "IVA": "44,00",
                "Importo fattura": "244,00",
                "Identificativo SDI": "0",
            },
            {
                "Creditore": "Ragione C",
                "Numero fattura": "F003",
                "Data emissione": "2025-01-12",
                "Data documento": "2025-01-12",
                "Data pagamento": "",
                "IVA": "0",
                "Importo fattura": "100",
                "Identificativo SDI": "123",
            },
        ]
    )

    processor = PisaRicevuteFTFileProcessor()
    input_path = tmp_path / "input_pisa_ricevute.xlsx"
    output_path = tmp_path / "output_pisa_ricevute.xlsx"
    df.to_excel(input_path, index=False)

    stats = processor.process_file(input_path, output_path)

    assert output_path.exists()
    assert stats["total_records"] == 3
    assert stats["fase2_records"] == 0
    assert stats["fase3_records"] == 3


def test_compare_files_all_period(tmp_path: Path):
    nfs_df = pd.DataFrame(
        {
            "C_NOME": ["A", "B"],
            "FAT_DATDOC": ["2025-01-05", "2025-01-15"],
            "FAT_NDOC": ["F001", "F002"],
            "DATA_REG_FATTURA": ["2025-01-10", "2025-01-20"],
            "FAT_PROT": ["P", "EP"],
            "FAT_NUM": [1, 2],
            "IMPONIBILE": [100.0, 200.0],
            "FAT_TOTFAT": [122.0, 244.0],
            "FAT_TOTIVA": [22.0, 44.0],
            "RA_IMPON": [100.0, 200.0],
            "RA_CODTRIB": ["I9", "RO"],
            "RA_IMPOSTA": [5.0, 10.0],
            "TMC_G8": ["", "ID2"],
        }
    )

    pisa_columns = [
        "Identificativo SDI",
        "B",
        "C",
        "D",
        "E",
        "F",
        "G",
        "Creditore",
        "I",
        "Importo Fattura",
        "K",
        "Importo Pagato",
        "M",
        "N",
        "O",
    ]
    pisa_df = pd.DataFrame(
        [
            ["", "b1", "c1", "d1", "e1", "2025-01-12", "g1", "Ragione A", "i1", 120.0, "k1", 150.0, "m1", "n1", "o1"],
            ["123", "b2", "c2", "d2", "e2", "2025-01-20", "g2", "Ragione B", "i2", 220.0, "k2", 250.0, "m2", "n2", "o2"],
            ["", "bX", "c_extra", "dX", "eX", "2025-01-25", "gX", "Ragione X", "iX", 50.0, "kX", 50.0, "mX", "nX", "oX"],
            ["", "b3", "c3", "d3", "e3", "2025-02-05", "g3", "Ragione C", "i3", 320.0, "k3", 300.0, "m3", "n3", "o3"],
        ],
        columns=pisa_columns,
    )

    nfs_path = tmp_path / "nfs.xlsx"
    pisa_path = tmp_path / "pisa.xlsx"
    output_path = tmp_path / "compare.xlsx"
    nfs_df.to_excel(nfs_path, index=False)
    pisa_df.to_excel(pisa_path, index=False)

    processor = CompareFTFileProcessor()
    summary = processor.process_files(nfs_path, pisa_path, output_path)

    assert output_path.exists()
    assert summary["period"] == "Tutto il periodo"
    assert summary["nfs"]["cartacee"]["count"] == 1
    assert summary["nfs"]["elettroniche"]["count"] == 1
    assert summary["pisa"]["cartacee"]["count"] == 3
    assert summary["pisa"]["elettroniche"]["count"] == 1

    wb = load_workbook(output_path, data_only=True)
    assert "Confronto" in wb.sheetnames
    assert "Differenze tra file" in wb.sheetnames
    assert "Dettaglio Diff. Cartacee" in wb.sheetnames
    assert "Dettaglio Diff. Elettroniche" in wb.sheetnames
    confronto_ws = wb["Confronto"]
    diff_ws = wb["Differenze tra file"]
    cart_ws = wb["Dettaglio Diff. Cartacee"]
    elet_ws = wb["Dettaglio Diff. Elettroniche"]

    assert diff_ws["A2"].value == "Cartacee"
    assert diff_ws["A3"].value == "Elettroniche"
    assert diff_ws["A4"].value == "Totale"

    assert float(diff_ws["G2"].value) == pytest.approx(float(confronto_ws["G2"].value), abs=0.01)
    assert float(diff_ws["G3"].value) == pytest.approx(float(confronto_ws["G3"].value), abs=0.01)
    assert float(diff_ws["G4"].value) == pytest.approx(float(confronto_ws["G4"].value), abs=0.01)

    assert float(cart_ws[f"L{cart_ws.max_row}"].value) == pytest.approx(float(confronto_ws["G2"].value), abs=0.01)
    assert int(elet_ws[f"G{elet_ws.max_row}"].value) == int(confronto_ws["F3"].value)
    assert float(elet_ws[f"H{elet_ws.max_row}"].value) == pytest.approx(float(confronto_ws["G3"].value), abs=0.01)


def test_compare_files_accepts_ragione_sociale_for_pisa(tmp_path: Path):
    nfs_df = pd.DataFrame(
        {
            "C_NOME": ["A"],
            "FAT_DATDOC": ["2025-01-05"],
            "FAT_NDOC": ["F001"],
            "DATA_REG_FATTURA": ["2025-01-10"],
            "FAT_PROT": ["EP"],
            "FAT_NUM": [1],
            "IMPONIBILE": [100.0],
            "FAT_TOTFAT": [122.0],
            "FAT_TOTIVA": [22.0],
            "RA_IMPON": [100.0],
            "RA_CODTRIB": [""],
            "RA_IMPOSTA": [0.0],
            "TMC_G8": ["123"],
        }
    )
    pisa_df = pd.DataFrame(
        {
            "Identificativo SDI": ["123"],
            "Numero fattura": ["F001"],
            "Data emissione": ["2025-01-12"],
            "Importo fattura": [122.0],
            "Ragione Sociale": ["A"],
        }
    )

    nfs_path = tmp_path / "nfs.xlsx"
    pisa_path = tmp_path / "pisa.xlsx"
    output_path = tmp_path / "compare.xlsx"
    nfs_df.to_excel(nfs_path, index=False)
    pisa_df.to_excel(pisa_path, index=False)

    summary = CompareFTFileProcessor().process_files(nfs_path, pisa_path, output_path)

    assert output_path.exists()
    assert summary["pisa"]["elettroniche"]["count"] == 1


def test_process_file_pisa_pagato_accepts_ragione_sociale(tmp_path: Path):
    df = pd.DataFrame(
        [
            {
                "Ragione Sociale": "Ragione A",
                "Numero fattura": "F001",
                "Identificativo SDI": "",
                "Data emissione": "2025-01-10",
                "Importo fattura": 120.0,
                "Importo pagato": 100.0,
            }
        ]
    )
    processor = PisaFTFileProcessor()
    input_path = tmp_path / "pisa_pagato.xlsx"
    output_path = tmp_path / "pisa_pagato_out.xlsx"
    df.to_excel(input_path, index=False)

    stats = processor.process_file(input_path, output_path)

    assert output_path.exists()
    assert stats["fase2_records"] == 1
    assert stats["fase3_records"] == 0
