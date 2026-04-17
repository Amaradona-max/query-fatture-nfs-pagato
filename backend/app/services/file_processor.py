from pathlib import Path
from typing import Any, Dict, List, Optional
import logging
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


logger = logging.getLogger(__name__)


class NFSFTFileProcessor:
    PROTOCOLLI_FASE2 = ["P", "2P", "L", "FCBI", "FCSI", "FCBE", "FCSE"]
    PROTOCOLLI_FASE3 = [
        "EP",
        "2EP",
        "EL",
        "2EL",
        "EZ",
        "2EZ",
        "EZP",
        "FPIC",
        "FSIC",
        "FPEC",
        "FSEC",
    ]
    PROTOCOLLI_FASE4 = ["AFIC", "ASIC", "AFEC", "ASEC", "ACBI", "ACSI", "ACBE", "ACSE"]

    DESCRIZIONI_FASE2 = {
        "P": "Fatture Cartacee San",
        "2P": "Fatture Cartacee Ter",
        "L": "Fatture Lib.Prof. San",
        "FCBI": "Fatture Cartacee Estere San",
        "FCSI": "Fatture Cartacee Estere San",
        "FCBE": "Fatture Cartacee Estere San",
        "FCSE": "Fatture Cartacee Estere San",
    }

    DESCRIZIONI_FASE3 = {
        "EP": "Fatture Elettroniche San",
        "2EP": "Fatture Elettroniche Ter",
        "EL": "Fatture Elettroniche Lib.Prof. San",
        "2EL": "Fatture Elettroniche Lib.Prof. Ter",
        "EZ": "Fatture Elettroniche Commerciali San",
        "2EZ": "Fatture Elettroniche Commerciali Ter",
        "EZP": "Fatture Elettroniche Commerciali San",
        "FPIC": "Fatture Elettroniche Estere San",
        "FSIC": "Fatture Elettroniche Estere San",
        "FPEC": "Fatture Elettroniche Estere San",
        "FSEC": "Fatture Elettroniche Estere San",
    }

    DESCRIZIONI_FASE4 = {
        "AFIC": "Autofatture Elettroniche San",
        "ASIC": "Autofatture Elettroniche San",
        "AFEC": "Autofatture Elettroniche San",
        "ASEC": "Autofatture Elettroniche San",
        "ACBI": "Autofatture Elettroniche San",
        "ACSI": "Autofatture Elettroniche San",
        "ACBE": "Autofatture Elettroniche San",
        "ACSE": "Autofatture Elettroniche San",
    }

    REQUIRED_COLUMNS = [
        "C_NOME",
        "FAT_DATDOC",
        "FAT_NDOC",
        "FAT_DATREG",
        "FAT_PROT",
        "FAT_NUM",
        "IMPONIBILE",
        "FAT_TOTIVA",
        "TMC_G8",
    ]

    OPTIONAL_COLUMNS_DEFAULTS: Dict[str, Any] = {
        "RA_IMPON": 0.0,
        "RA_CODTRIB": "",
        "RA_IMPOSTA": 0.0,
        "DMA_NUM": "",
        "TMA_TOT": 0.0,
        "IMPORTO_PAGATO": 0.0,
        "DATA_ESEGUITO_BANCARIO": "",
        "DATA_GEN_MANDATO": "",
        "M2_TMC_DATREG": "",
        "FAT_TOTFAT": 0.0,
    }

    def __init__(self) -> None:
        self.all_protocols = self.PROTOCOLLI_FASE2 + self.PROTOCOLLI_FASE3 + self.PROTOCOLLI_FASE4
        self._last_csv_skipped_lines = 0

    def _read_csv(self, input_path: Path, **kwargs) -> pd.DataFrame:
        last_error: Optional[Exception] = None
        self._last_csv_skipped_lines = 0
        for encoding in ("utf-8-sig", "latin-1"):
            try:
                if "on_bad_lines" not in kwargs:
                    return pd.read_csv(
                        input_path,
                        sep=None,
                        engine="python",
                        encoding=encoding,
                        on_bad_lines="error",
                        **kwargs,
                    )
                return pd.read_csv(input_path, sep=None, engine="python", encoding=encoding, **kwargs)
            except pd.errors.ParserError as exc:
                try:
                    logger.warning("CSV malformato, salto righe non parseabili: %s", exc)
                    df = pd.read_csv(
                        input_path,
                        sep=None,
                        engine="python",
                        encoding=encoding,
                        on_bad_lines="skip",
                        **kwargs,
                    )
                    return df
                except Exception as exc2:
                    last_error = exc2
            except Exception as exc:
                last_error = exc
        if last_error:
            raise last_error
        raise ValueError("Impossibile leggere il file CSV")

    def _read_tabular_file(self, input_path: Path, **kwargs) -> pd.DataFrame:
        if input_path.suffix.lower() == ".csv":
            return self._read_csv(input_path, **kwargs)
        return pd.read_excel(input_path, **kwargs)

    def validate_file(self, df: pd.DataFrame) -> None:
        def normalize_col_name(value: Any) -> str:
            text = str(value).strip().upper()
            return re.sub(r"[^A-Z0-9]", "", text)

        df.columns = [str(c).strip() for c in df.columns]
        normalized_to_original = {normalize_col_name(c): c for c in df.columns}

        # Supporta tracciati alternativi (es. NFS Pagato I° Trim. 2026)
        alias_map = {
            "RAGIONESOCIALE": "C_NOME",
            "FTPROT": "FAT_PROT",
            "NPROTOCOLLO": "FAT_NUM",
            "NUMEROMANDATO": "FAT_NUM",
            "NFATTURA": "FAT_NUM",
            "NFATTURE": "FAT_NDOC",
            "NDOCUMENTO": "FAT_NDOC",
            "DATAFATTURA": "FAT_DATDOC",
            "IMPTOTIVA": "FAT_TOTIVA",
            "IMPTOTFATTURA": "FAT_TOTFAT",
            "IDENTSDI": "TMC_G8",
            "IMPORTOPAGATO": "IMPORTO_PAGATO",
            "VALUTAIMPORTOMANDATO": "M2_TMC_DATREG",
            "DMANUM": "DMA_NUM",
            "NUMMANDATO": "DMA_NUM",
            "TOTIMPORTOMANDATO": "TMA_TOT",
            "DATAESEGUITOBANCARIO": "DATA_ESEGUITO_BANCARIO",
            "DATAGENMANDATO": "DATA_GEN_MANDATO",
            "CODTRIBUTO": "RA_CODTRIB",
            "IMPTOTRIT": "RA_IMPOSTA",
            "IMPONIBILERITENUTADACC": "RA_IMPON",
        }
        rename_by_alias: Dict[str, str] = {}
        assigned_targets = set(df.columns)
        for original_col in df.columns:
            mapped = alias_map.get(normalize_col_name(original_col))
            if mapped and mapped not in assigned_targets:
                rename_by_alias[original_col] = mapped
                assigned_targets.add(mapped)
        if rename_by_alias:
            df.rename(columns=rename_by_alias, inplace=True)
            df.columns = [str(c).strip() for c in df.columns]
            normalized_to_original = {normalize_col_name(c): c for c in df.columns}

        for canonical in list(self.REQUIRED_COLUMNS) + list(self.OPTIONAL_COLUMNS_DEFAULTS.keys()):
            if canonical in df.columns:
                continue
            key = normalize_col_name(canonical)
            original = normalized_to_original.get(key)
            if original and original in df.columns:
                df.rename(columns={original: canonical}, inplace=True)

        if "FAT_DATREG" not in df.columns:
            for alt in ("DATA_REG_FATTURA", "FAT_REG_FATTURA", "DATAREGFATTURA", "FATREGFATTURA", "DATAREGISTRAZIONE"):
                original = normalized_to_original.get(normalize_col_name(alt))
                if original and original in df.columns:
                    df.rename(columns={original: "FAT_DATREG"}, inplace=True)
                    break

        if "FAT_NUM" not in df.columns and "FAT_NDOC" in df.columns:
            df["FAT_NUM"] = df["FAT_NDOC"]
        if "FAT_NDOC" not in df.columns and "FAT_NUM" in df.columns:
            df["FAT_NDOC"] = df["FAT_NUM"]

        missing_cols = [col for col in self.REQUIRED_COLUMNS if col not in df.columns]
        if missing_cols:
            raise ValueError(f"Colonne mancanti: {', '.join(missing_cols)}")

        for col, default in self.OPTIONAL_COLUMNS_DEFAULTS.items():
            if col not in df.columns:
                df[col] = default

    def _read_excel_flexible(self, input_path: Path) -> pd.DataFrame:
        if input_path.suffix.lower() == ".csv":
            try:
                df = self._read_csv(input_path)
                if len(df.columns) > 0:
                    df.columns = [str(c).strip() for c in df.columns]
                    has_real_headers = any(col and not col.lower().startswith("unnamed") for col in df.columns)
                    if has_real_headers:
                        return df
            except Exception:
                pass

            try:
                raw = self._read_csv(input_path, header=None, nrows=25)
            except Exception:
                return self._read_csv(input_path)

            wanted = set(self.REQUIRED_COLUMNS) | set(self.OPTIONAL_COLUMNS_DEFAULTS.keys()) | {"DATA_REG_FATTURA", "FAT_REG_FATTURA"}
            wanted_upper = {str(x).strip().upper() for x in wanted}

            header_row_idx: Optional[int] = None
            for idx in range(len(raw)):
                values = raw.iloc[idx].tolist()
                normalized = {str(v).strip().upper() for v in values if v is not None and str(v).strip() != ""}
                if len(normalized & wanted_upper) >= 5:
                    header_row_idx = idx
                    break

            if header_row_idx is None:
                return self._read_csv(input_path)

            df = self._read_csv(input_path, header=header_row_idx)
            df.columns = [str(c).strip() for c in df.columns]
            return df

        try:
            df = pd.read_excel(input_path)
            if len(df.columns) > 0:
                df.columns = [str(c).strip() for c in df.columns]
                has_real_headers = any(col and not col.lower().startswith("unnamed") for col in df.columns)
                if has_real_headers:
                    return df
        except Exception:
            pass

        try:
            raw = pd.read_excel(input_path, header=None, nrows=25)
        except Exception:
            return pd.read_excel(input_path)

        wanted = set(self.REQUIRED_COLUMNS) | set(self.OPTIONAL_COLUMNS_DEFAULTS.keys()) | {"DATA_REG_FATTURA", "FAT_REG_FATTURA"}
        wanted_upper = {str(x).strip().upper() for x in wanted}

        header_row_idx: Optional[int] = None
        for idx in range(len(raw)):
            values = raw.iloc[idx].tolist()
            normalized = {str(v).strip().upper() for v in values if v is not None and str(v).strip() != ""}
            if len(normalized & wanted_upper) >= 5:
                header_row_idx = idx
                break

        if header_row_idx is None:
            return pd.read_excel(input_path)

        df = pd.read_excel(input_path, header=header_row_idx)
        df.columns = [str(c).strip() for c in df.columns]
        return df

    def _parse_mixed_date_series(self, series: pd.Series) -> pd.Series:
        if pd.api.types.is_datetime64_any_dtype(series):
            return series
        as_text = series.astype(str).str.strip()
        iso_mask = as_text.str.match(r"^\d{4}-\d{2}-\d{2}( \d{2}:\d{2}:\d{2})?$")
        parsed_iso = pd.to_datetime(series.where(iso_mask), errors="coerce", dayfirst=False)
        parsed_other = pd.to_datetime(series.where(~iso_mask), errors="coerce", dayfirst=True)
        return parsed_iso.fillna(parsed_other)

    def _extract_quarter_period(self, input_path: Path) -> Optional[tuple[pd.Timestamp, pd.Timestamp]]:
        name = input_path.name.upper()
        match = re.search(r"(I{1,3}|IV|[1-4])\s*°?\s*TRIM(?:ESTRE)?\.?\s*[-_ ]*(20\d{2})", name)
        if not match:
            return None
        quarter_raw, year_raw = match.groups()
        quarter_map = {"I": 1, "II": 2, "III": 3, "IV": 4}
        quarter = quarter_map[quarter_raw] if quarter_raw in quarter_map else int(quarter_raw)
        year = int(year_raw)
        month_start = (quarter - 1) * 3 + 1
        start = pd.Timestamp(year=year, month=month_start, day=1)
        end = start + pd.offsets.QuarterEnd()
        return start, end

    def _extract_quarter_period_extended(self, input_path: Path) -> Optional[tuple[pd.Timestamp, pd.Timestamp]]:
        period = self._extract_quarter_period(input_path)
        if period is None:
            return None
        start, end = period
        return start, end + pd.Timedelta(days=8)

    def _extract_quarter_period_extended(self, input_path: Path) -> Optional[tuple[pd.Timestamp, pd.Timestamp]]:
        period = self._extract_quarter_period(input_path)
        if period is None:
            return None
        start, end = period
        return start, end + pd.Timedelta(days=8)

    def _extract_quarter_period_extended(self, input_path: Path) -> Optional[tuple[pd.Timestamp, pd.Timestamp]]:
        period = self._extract_quarter_period(input_path)
        if period is None:
            return None
        start, end = period
        return start, end + pd.Timedelta(days=8)

    def _filter_by_file_quarter_extended(self, df: pd.DataFrame, date_column: str, input_path: Path) -> pd.DataFrame:
        if date_column not in df.columns:
            return df
        period = self._extract_quarter_period_extended(input_path)
        if period is None:
            return df
        start, end = period
        date_series = self._parse_mixed_date_series(df[date_column])
        return df[date_series.between(start, end)].copy()

    def _filter_by_file_quarter(self, df: pd.DataFrame, date_column: str, input_path: Path) -> pd.DataFrame:
        if date_column not in df.columns:
            return df
        period = self._extract_quarter_period(input_path)
        if period is None:
            return df
        start, end = period
        date_series = self._parse_mixed_date_series(df[date_column])
        mask = date_series.between(start, end)
        return df[mask].copy()

    def _split_by_sdi(self, df: pd.DataFrame, sdi_column: str) -> tuple[pd.DataFrame, pd.DataFrame]:
        sdi_series = df[sdi_column]
        normalized = sdi_series.astype(str).str.strip().where(~sdi_series.isna(), "")
        normalized = normalized.str.lower().str.replace(",", ".", regex=False)
        empty_text_mask = normalized.isin(["", "nan", "none", "null"])
        numeric = pd.to_numeric(normalized, errors="coerce")
        zero_mask = numeric.eq(0) & ~numeric.isna()
        empty_mask = empty_text_mask | zero_mask
        cartacee_df = df[empty_mask].copy()
        elettroniche_df = df[~empty_mask].copy()
        return cartacee_df, elettroniche_df

    def process_file(self, input_path: Path, output_path: Path) -> Dict[str, Any]:
        try:
            logger.info("Caricamento file: %s", input_path)
            df = self._read_excel_flexible(input_path)
            df.columns = [str(c).strip() for c in df.columns]

            self.validate_file(df)

            df["FAT_PROT"] = df["FAT_PROT"].astype(str).str.strip().str.upper()
            totale_iniziale = len(df)
            df_senza_duplicati = df.drop_duplicates(subset=["FAT_NDOC", "FAT_DATDOC", "C_NOME"]).copy()
            duplicati_rimossi = totale_iniziale - len(df_senza_duplicati)
            df_filtrato = df_senza_duplicati[df_senza_duplicati["FAT_PROT"].isin(self.all_protocols)].copy()

            if len(df_filtrato) == 0:
                raise ValueError("Nessun protocollo valido trovato nel file")

            df_filtrato["RA_CODTRIB"] = df_filtrato["RA_CODTRIB"].astype(str).str.strip().str.upper()

            colonne_ordinate = [
                "C_NOME",
                "FAT_DATDOC",
                "FAT_NDOC",
                "FAT_DATREG",
                "FAT_PROT",
                "FAT_NUM",
                "IMPONIBILE",
                "FAT_TOTIVA",
                "RA_CODTRIB",
                "DMA_NUM",
                "TMA_TOT",
                "TMC_G8",
                "M2_TMC_DATREG",
            ]

            df_finale = df_filtrato[colonne_ordinate].copy()
            df_finale.columns = [
                "Ragione Sociale",
                "Data Fatture",
                "N. Fatture",
                "Data Ricevimento",
                "Protocollo",
                "N. Protocollo",
                "Tot. Imponibile",
                "Imposta",
                "Codice Tributo",
                "N. Mandato",
                "Tot. Importo Mandato",
                "Id. SDI",
                "Valuta Importo Mandato",
            ]

            df_finale["Data Fatture"] = pd.to_datetime(df_finale["Data Fatture"], errors="coerce")
            df_finale["Data Ricevimento"] = pd.to_datetime(df_finale["Data Ricevimento"], errors="coerce")
            df_finale["Tot. Imponibile"] = pd.to_numeric(
                df_finale["Tot. Imponibile"].astype(str).str.replace(".", "", regex=False).str.replace(",", ".", regex=False),
                errors="coerce",
            ).fillna(0)
            df_finale["Importo Pagato"] = df_finale["Tot. Imponibile"]

            df_finale = df_finale.sort_values("Data Ricevimento")

            df_dati = df_finale.copy()
            ordered_columns = [
                "Ragione Sociale",
                "Data Fatture",
                "N. Fatture",
                "Data Ricevimento",
                "Protocollo",
                "N. Protocollo",
                "Tot. Imponibile",
                "Importo Pagato",
                "Imposta",
                "Codice Tributo",
                "N. Mandato",
                "Tot. Importo Mandato",
                "Id. SDI",
                "Valuta Importo Mandato",
            ]
            df_dati = df_dati[[col for col in ordered_columns if col in df_dati.columns]]

            stats = self._calculate_stats(df_finale, duplicati_rimossi)
            self._create_excel_output(df_finale, output_path, display_df=df_dati)

            logger.info("File elaborato con successo: %s", stats)
            return stats
        except Exception as exc:
            logger.error("Errore elaborazione file: %s", str(exc))
            raise

    def _calculate_stats(self, df: pd.DataFrame, duplicates_removed: int) -> Dict[str, Any]:
        cartacee_df = df[df["Protocollo"].astype(str).str.upper().isin(self.PROTOCOLLI_FASE2)].copy()
        elettroniche_df = df[df["Protocollo"].astype(str).str.upper().isin(self.PROTOCOLLI_FASE3 + self.PROTOCOLLI_FASE4)].copy()
        fase2_count = int(len(cartacee_df))
        fase3_count = int(len(elettroniche_df))
        fase2_amount = round(float(pd.to_numeric(cartacee_df["Importo Pagato"], errors="coerce").fillna(0).sum()), 2)
        fase3_amount = round(float(pd.to_numeric(elettroniche_df["Importo Pagato"], errors="coerce").fillna(0).sum()), 2)
        protocols_fase2 = {"Cartacee": fase2_count}
        protocols_fase3 = {"Elettroniche": fase3_count}

        return {
            "total_records": len(df),
            "fase2_records": fase2_count,
            "fase3_records": fase3_count,
            "fase2_amount": fase2_amount,
            "fase3_amount": fase3_amount,
            "duplicates_removed": duplicates_removed,
            "protocols_fase2": protocols_fase2,
            "protocols_fase3": protocols_fase3,
        }

    def _count_by_protocol(self, df: pd.DataFrame, protocols: list) -> Dict[str, int]:
        counts = {}
        for prot in protocols:
            counts[prot] = len(df[df["Protocollo"] == prot])
        return counts

    def _create_excel_output(
        self,
        df: pd.DataFrame,
        output_path: Path,
        display_df: Optional[pd.DataFrame] = None,
    ) -> None:
        wb = Workbook()

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        total_font = Font(bold=True)

        ws_dati = self._add_dataframe_sheet(
            wb,
            "Dati",
            display_df if display_df is not None else df,
            header_fill,
            header_font,
            total_fill,
            total_font,
            date_columns=["Data Fatture", "Data Ricevimento", "Valuta Importo Mandato"],
            money_columns=["Tot. Imponibile", "Importo Pagato", "Imposta", "Tot. Importo Mandato"],
            use_active=True,
        )

        cartacee_df = df[df["Protocollo"].astype(str).str.upper().isin(self.PROTOCOLLI_FASE2)].copy()
        elettroniche_df = df[df["Protocollo"].astype(str).str.upper().isin(self.PROTOCOLLI_FASE3)].copy()
        autofatture_df = df[df["Protocollo"].astype(str).str.upper().isin(self.PROTOCOLLI_FASE4)].copy()

        ws_nota2 = wb.create_sheet("Fatture Cartacee")
        self._create_summary_sheet(
            ws_nota2,
            cartacee_df,
            self.PROTOCOLLI_FASE2,
            self.DESCRIZIONI_FASE2,
            header_fill,
            header_font,
            total_fill,
            total_font,
        )

        ws_nota3 = wb.create_sheet("Fatture Elettroniche")
        self._create_summary_sheet(
            ws_nota3,
            elettroniche_df,
            self.PROTOCOLLI_FASE3,
            self.DESCRIZIONI_FASE3,
            header_fill,
            header_font,
            total_fill,
            total_font,
        )

        ws_nota4 = wb.create_sheet("Autofatture Elettroniche Estere")
        self._create_summary_sheet(
            ws_nota4,
            autofatture_df,
            self.PROTOCOLLI_FASE4,
            self.DESCRIZIONI_FASE4,
            header_fill,
            header_font,
            total_fill,
            total_font,
        )

        wb.save(output_path)

    def _add_dataframe_sheet(
        self,
        wb: Workbook,
        title: str,
        df: pd.DataFrame,
        header_fill: PatternFill,
        header_font: Font,
        total_fill: PatternFill,
        total_font: Font,
        date_columns=None,
        money_columns=None,
        date_format: str = "mm/dd/yyyy",
        add_totals: bool = True,
        auto_size: bool = True,
        use_active: bool = False,
    ):
        ws = wb.active if use_active else wb.create_sheet(title)
        ws.title = title

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        if auto_size:
            sample_rows = 50
            max_row = min(ws.max_row, sample_rows + 1)
            for column in ws.iter_cols(max_row=max_row):
                max_len = max((len(str(c.value or "")) for c in column), default=8)
                letter = column[0].column_letter
                ws.column_dimensions[letter].width = min(max_len + 2, 45)

        date_columns = date_columns or []
        money_columns = money_columns or []
        money_columns = [column for column in money_columns if column in df.columns]
        header_index = {cell.value: cell.column for cell in ws[1]}
        money_format = "#,##0.00"

        for column_name in date_columns:
            column_index = header_index.get(column_name)
            if column_index:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=column_index, max_col=column_index):
                    for cell in row:
                        if cell.value is not None:
                            cell.number_format = date_format

        for column_name in money_columns:
            column_index = header_index.get(column_name)
            if column_index:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=column_index, max_col=column_index):
                    for cell in row:
                        if cell.value is not None:
                            cell.number_format = money_format

        if money_columns and add_totals:
            totals = {}
            for column_name in money_columns:
                totals[column_name] = pd.to_numeric(df[column_name], errors="coerce").sum()
            total_row = ["TOTALE"] + [""] * (len(df.columns) - 1)
            for column_name, total_value in totals.items():
                total_row[df.columns.get_loc(column_name)] = total_value
            ws.append(total_row)
            total_row_index = ws.max_row
            for cell in ws[total_row_index]:
                cell.fill = total_fill
                cell.font = total_font
            for column_name in money_columns:
                column_index = header_index.get(column_name)
                if column_index:
                    ws.cell(row=total_row_index, column=column_index).number_format = money_format

        return ws

    def _create_summary_sheet(self, ws, df, protocols, descriptions, header_fill, header_font, total_fill, total_font):
        ws["A1"] = "PROTOCOLLO"
        ws["B1"] = "DESCRIZIONE"
        ws["C1"] = "NUMERO TOTALE"
        ws["D1"] = "IMPORTO PAGATO"

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        money_format = "#,##0.00"
        row = 2
        for prot in protocols:
            count = len(df[df["Protocollo"] == prot])
            imponibile_totale = pd.to_numeric(df.loc[df["Protocollo"] == prot, "Importo Pagato"], errors="coerce").fillna(0).sum()
            ws[f"A{row}"] = prot
            ws[f"B{row}"] = descriptions[prot]
            ws[f"C{row}"] = count
            ws[f"D{row}"] = imponibile_totale
            ws[f"D{row}"].number_format = money_format
            row += 1

        ws[f"A{row}"] = "TOTALE"
        ws[f"A{row}"].font = total_font
        ws[f"C{row}"] = f"=SUM(C2:C{row - 1})"
        ws[f"C{row}"].font = total_font
        ws[f"D{row}"] = f"=SUM(D2:D{row - 1})"
        ws[f"D{row}"].number_format = money_format
        for cell in ws[row]:
            cell.fill = total_fill
            cell.font = total_font

        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20


class PisaFTFileProcessor(NFSFTFileProcessor):
    MAX_DETAIL_ROWS = 5000

    def _to_number_series_pisa(self, series: pd.Series) -> pd.Series:
        normalized = series.astype(str).str.replace(" ", "", regex=False).str.strip()
        has_dot = normalized.str.contains(r"\.", regex=True, na=False)
        has_comma = normalized.str.contains(",", regex=False, na=False)
        both_mask = has_dot & has_comma
        normalized = normalized.where(~both_mask, normalized.str.replace(".", "", regex=False))
        normalized = normalized.str.replace(",", ".", regex=False)
        return pd.to_numeric(normalized, errors="coerce").fillna(0)

    def process_file(self, input_path: Path, output_path: Path) -> Dict[str, Any]:
        try:
            logger.info("Caricamento file Pisa Pagato: %s", input_path)
            df = pd.read_excel(input_path, dtype=str)
            df.columns = [str(c).strip() for c in df.columns]

            def pick_column(*names: str) -> Optional[str]:
                lower_map = {str(c).strip().lower(): c for c in df.columns}
                for n in names:
                    key = n.strip().lower()
                    if key in lower_map:
                        return lower_map[key]
                return None

            sdi_column = pick_column("Identificativo SDI")
            creditore_col = pick_column("Creditore")
            numero_fattura_col = pick_column("Numero fattura")
            data_emissione_col = pick_column("Data emissione")
            data_pagamento_col = pick_column("Data pagamento")

            amount_col = (
                pick_column("Importo pagato")
                or pick_column("Importo liquidato")
                or pick_column("Importo fattura")
            )

            missing = [label for label, col in [
                ("Identificativo SDI", sdi_column),
                ("Creditore", creditore_col),
                ("Numero fattura", numero_fattura_col),
                ("Data emissione", data_emissione_col),
                ("Data pagamento", data_pagamento_col),
                ("Importo pagato/liquidato/fattura", amount_col),
            ] if col is None]
            if missing:
                raise ValueError(f"Colonne mancanti nel file Pisa: {', '.join(missing)}")

            pagamento_series = df[data_pagamento_col]
            pagamento_mask = ~(pagamento_series.isna() | (pagamento_series.astype(str).str.strip() == ""))
            df_pagato = df[pagamento_mask].copy()

            df_pagato = self._filter_by_file_quarter(df_pagato, data_pagamento_col, input_path)

            df_finale = pd.DataFrame(
                {
                    "Creditore": df_pagato[creditore_col],
                    "Numero fattura": df_pagato[numero_fattura_col],
                    "Data emissione": df_pagato[data_emissione_col],
                    "Data pagamento": df_pagato[data_pagamento_col],
                    "Importo pagato": df_pagato[amount_col],
                    "Identificativo SDI": df_pagato[sdi_column],
                }
            )

            cartacee_df, elettroniche_df = self._split_by_sdi(df_finale, "Identificativo SDI")
            df_dati = df_finale.copy()
            if len(df_dati) > self.MAX_DETAIL_ROWS:
                df_dati = df_dati.head(self.MAX_DETAIL_ROWS).copy()
            self._create_excel_output(df_finale, cartacee_df, elettroniche_df, output_path, display_df=df_dati)
            fase2_amount = round(
                float(self._to_number_series_pisa(cartacee_df["Importo pagato"]).sum()),
                2,
            )
            fase3_amount = round(
                float(self._to_number_series_pisa(elettroniche_df["Importo pagato"]).sum()),
                2,
            )
            stats = {
                "total_records": len(df_finale),
                "fase2_records": len(cartacee_df),
                "fase3_records": len(elettroniche_df),
                "fase2_amount": fase2_amount,
                "fase3_amount": fase3_amount,
                "duplicates_removed": 0,
                "protocols_fase2": {"Cartacee": len(cartacee_df)},
                "protocols_fase3": {"Elettroniche": len(elettroniche_df)},
            }
            logger.info("File Pisa Pagato elaborato con successo: %s", stats)
            return stats
        except Exception as exc:
            logger.error("Errore elaborazione file Pisa Pagato: %s", str(exc))
            raise

    def _create_excel_output(
        self,
        df: pd.DataFrame,
        cartacee_df: pd.DataFrame,
        elettroniche_df: pd.DataFrame,
        output_path: Path,
        display_df: Optional[pd.DataFrame] = None,
    ) -> None:
        wb = Workbook()

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        total_font = Font(bold=True)

        dati_df = display_df if display_df is not None else df
        self._add_dataframe_sheet(
            wb,
            "Dati",
            dati_df,
            header_fill,
            header_font,
            total_fill,
            total_font,
            date_columns=[column for column in dati_df.columns if "data" in str(column).lower()],
            date_format="dd/mm/yyyy",
            money_columns=[column for column in ("Importo pagato", "Importo fattura", "Importo liquidato", "Imponibile") if column in dati_df.columns],
            auto_size=False,
            use_active=True,
        )

        ws_cartacee = wb.create_sheet("Fatture Cartacee")
        self._create_simple_summary_sheet(
            ws_cartacee,
            cartacee_df,
            header_fill,
            header_font,
            total_fill,
            total_font,
        )

        ws_elettroniche = wb.create_sheet("Fatture Elettroniche")
        self._create_simple_summary_sheet(
            ws_elettroniche,
            elettroniche_df,
            header_fill,
            header_font,
            total_fill,
            total_font,
        )

        wb.save(output_path)

    def _split_by_sdi(self, df: pd.DataFrame, sdi_column: str) -> tuple[pd.DataFrame, pd.DataFrame]:
        sdi_series = df[sdi_column]
        normalized = sdi_series.astype(str).str.strip().where(~sdi_series.isna(), "")
        normalized = normalized.str.lower().str.replace(",", ".", regex=False)
        empty_text_mask = normalized.isin(["", "nan", "none", "null"])
        empty_mask = empty_text_mask
        cartacee_df = df[empty_mask].copy()
        elettroniche_df = df[~empty_mask].copy()
        return cartacee_df, elettroniche_df

    def _letters_to_indices(self, letters: list[str]) -> list[int]:
        return [ord(letter) - ord("A") for letter in letters]

    def _create_simple_summary_sheet(
        self,
        ws,
        df: pd.DataFrame,
        header_fill: PatternFill,
        header_font: Font,
        total_fill: PatternFill,
        total_font: Font,
    ) -> None:
        ws["A1"] = "NUMERO TOTALE"
        ws["B1"] = "IMPORTO PAGATO"

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        amount_col = "Importo pagato" if "Importo pagato" in df.columns else "Imponibile"
        imponibile_totale = float(self._to_number_series_pisa(df[amount_col]).sum())
        ws["A2"] = len(df)
        ws["B2"] = imponibile_totale
        ws["B2"].number_format = "#,##0.00"

        for cell in ws[2]:
            cell.fill = total_fill
            cell.font = total_font

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20


class PisaRicevuteFTFileProcessor(NFSFTFileProcessor):
    PHASE = 1
    INPUT_REQUIRED_COLUMNS = [
        "Creditore",
        "Numero fattura",
        "Data emissione",
        "Data documento",
        "Data pagamento",
        "IVA",
        "Importo fattura",
        "Identificativo SDI",
    ]
    OUTPUT_COLUMNS = [
        "Ragione sociale",
        "N.fatture",
        "Data emissione",
        "Data documento",
        "Data pagamento",
        "Ivam",
        "Imponibile",
        "Totale fatture",
        "Identificativo SDI",
    ]
    OUTPUT_DATE_COLUMNS = ["Data emissione", "Data documento", "Data pagamento"]
    OUTPUT_MONEY_COLUMNS = ["Ivam", "Imponibile", "Totale fatture"]
    MAX_DETAIL_ROWS = 5000

    def process_file(self, input_path: Path, output_path: Path) -> Dict[str, Any]:
        try:
            logger.info("Caricamento file Pisa Ricevute: %s", input_path)
            try:
                df = self._read_tabular_file(input_path, usecols=self.INPUT_REQUIRED_COLUMNS, dtype=str)
            except ValueError:
                df_header = self._read_tabular_file(input_path, nrows=0)
                missing_columns = [col for col in self.INPUT_REQUIRED_COLUMNS if col not in df_header.columns]
                if missing_columns:
                    raise ValueError(f"Colonne mancanti: {', '.join(missing_columns)}")
                raise

            totale_fattura = pd.to_numeric(
                df["Importo fattura"].astype(str).str.replace(",", ".", regex=False),
                errors="coerce",
            ).fillna(0)
            iva = pd.to_numeric(
                df["IVA"].astype(str).str.replace(",", ".", regex=False),
                errors="coerce",
            ).fillna(0)
            imponibile = totale_fattura - iva

            df_finale = pd.DataFrame(
                {
                    "Ragione sociale": df["Creditore"],
                    "N.fatture": df["Numero fattura"],
                    "Data emissione": pd.to_datetime(df["Data emissione"], errors="coerce"),
                    "Data documento": pd.to_datetime(df["Data documento"], errors="coerce"),
                    "Data pagamento": pd.to_datetime(df["Data pagamento"], errors="coerce"),
                    "Ivam": iva,
                    "Imponibile": imponibile,
                    "Totale fatture": totale_fattura,
                    "Identificativo SDI": df["Identificativo SDI"],
                }
            )
            df_finale = df_finale[self.OUTPUT_COLUMNS]
            df_finale = self._filter_by_file_quarter(df_finale, "Data pagamento", input_path)

            cartacee_df, elettroniche_df = self._split_by_sdi(df_finale, "Identificativo SDI")
            display_df = df_finale
            if len(display_df) > self.MAX_DETAIL_ROWS:
                display_df = display_df.head(self.MAX_DETAIL_ROWS).copy()
            self._create_excel_output(df_finale, cartacee_df, elettroniche_df, output_path, display_df=display_df)
            fase2_amount = round(float(pd.to_numeric(cartacee_df["Totale fatture"], errors="coerce").fillna(0).sum()), 2)
            fase3_amount = round(float(pd.to_numeric(elettroniche_df["Totale fatture"], errors="coerce").fillna(0).sum()), 2)
            stats = {
                "total_records": len(df_finale),
                "fase2_records": len(cartacee_df),
                "fase3_records": len(elettroniche_df),
                "fase2_amount": fase2_amount,
                "fase3_amount": fase3_amount,
                "duplicates_removed": 0,
                "protocols_fase2": {"Cartacee": len(cartacee_df)},
                "protocols_fase3": {"Elettroniche": len(elettroniche_df)},
            }
            logger.info("File Pisa Ricevute elaborato con successo: %s", stats)
            return stats
        except Exception as exc:
            logger.error("Errore elaborazione file Pisa Ricevute: %s", str(exc))
            raise

    def _create_excel_output(
        self,
        df: pd.DataFrame,
        cartacee_df: pd.DataFrame,
        elettroniche_df: pd.DataFrame,
        output_path: Path,
        display_df: Optional[pd.DataFrame] = None,
    ) -> None:
        wb = Workbook()

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        total_font = Font(bold=True)

        dati_df = display_df if display_df is not None else df
        self._add_dataframe_sheet(
            wb,
            "Dati",
            dati_df,
            header_fill,
            header_font,
            total_fill,
            total_font,
            date_columns=self.OUTPUT_DATE_COLUMNS,
            date_format="dd/mm/yyyy",
            money_columns=self.OUTPUT_MONEY_COLUMNS,
            auto_size=False,
            use_active=True,
        )

        ws_cartacee = wb.create_sheet("Fatture Cartacee")
        self._create_simple_summary_sheet(
            ws_cartacee,
            cartacee_df,
            header_fill,
            header_font,
            total_fill,
            total_font,
        )

        ws_elettroniche = wb.create_sheet("Fatture Elettroniche")
        self._create_simple_summary_sheet(
            ws_elettroniche,
            elettroniche_df,
            header_fill,
            header_font,
            total_fill,
            total_font,
        )

        wb.save(output_path)

    def _split_by_sdi(self, df: pd.DataFrame, sdi_column: str) -> tuple[pd.DataFrame, pd.DataFrame]:
        sdi_series = df[sdi_column]
        normalized = sdi_series.astype(str).str.strip().where(~sdi_series.isna(), "")
        normalized = normalized.str.lower().str.replace(",", ".", regex=False)
        empty_text_mask = normalized.isin(["", "nan", "none", "null"])
        empty_mask = empty_text_mask
        cartacee_df = df[empty_mask].copy()
        elettroniche_df = df[~empty_mask].copy()
        return cartacee_df, elettroniche_df

    def _create_simple_summary_sheet(
        self,
        ws,
        df: pd.DataFrame,
        header_fill: PatternFill,
        header_font: Font,
        total_fill: PatternFill,
        total_font: Font,
    ) -> None:
        ws["A1"] = "NUMERO TOTALE"
        ws["B1"] = "TOTALE FATTURE"

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        totale_fatture = pd.to_numeric(df["Totale fatture"], errors="coerce").sum()
        ws["A2"] = len(df)
        ws["B2"] = totale_fatture
        ws["B2"].number_format = "#,##0.00"

        for cell in ws[2]:
            cell.fill = total_fill
            cell.font = total_font

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20


class CompareFTFileProcessor:
    NFS_REQUIRED_COLUMNS = [
        "C_NOME",
        "FAT_PROT",
        "FAT_NUM",
        "FAT_NDOC",
        "FAT_DATDOC",
        "FAT_DATREG",
        "M2_TMC_DATREG",
        "FAT_TOTIVA",
        "IMPONIBILE",
        "RA_CODTRIB",
        "TMC_G8",
    ]
    NFS_RENAME_MAP = {
        "C_NOME": "Ragione sociale",
        "FAT_PROT": "Prot.",
        "FAT_NUM": "FAT_NUM",
        "FAT_NDOC": "N.fatture",
        "FAT_DATDOC": "Data Fatture",
        "FAT_DATREG": "Datat reg.",
        "M2_TMC_DATREG": "Valuta Importo Mandato",
        "FAT_TOTIVA": "Iva",
        "IMPONIBILE": "Imponibile",
        "RA_CODTRIB": "Codice tributo",
        "TMC_G8": "Identificativo SDI",
    }
    NFS_CARTACEE_PROTOCOLS = {"P", "2P", "L", "FCBI", "FCSI", "FCBE", "FCSE"}
    NFS_ELETTRONICHE_PROTOCOLS = {"EP", "2EP", "EL", "2EL", "EZ", "2EZ", "EZP", "FPIC", "FSIC", "FPEC", "FSEC"}
    NFS_AUTOFATTURE_PROTOCOLS = {"AFIC", "ASIC", "AFEC", "ASEC", "ACBI", "ACSI", "ACBE", "ACSE"}
    NFS_ALLOWED_PROTOCOLS = NFS_CARTACEE_PROTOCOLS | NFS_ELETTRONICHE_PROTOCOLS | NFS_AUTOFATTURE_PROTOCOLS

    PISA_REQUIRED_COLUMNS = ["Creditore", "Numero fattura", "Identificativo SDI", "Data emissione", "Importo fattura"]
    NFS_OPTIONAL_DEFAULTS: Dict[str, Any] = {
        "RA_IMPON": 0.0,
        "RA_IMPOSTA": 0.0,
        "RA_CODTRIB": "",
        "M2_TMC_DATREG": "",
        "DMA_NUM": "",
        "TMA_TOT": 0.0,
        "DATA_GEN_MANDATO": "",
        "IMPORTO_PAGATO": 0.0,
    }

    def _normalize_col_name(self, value: Any) -> str:
        text = str(value).strip().upper()
        return re.sub(r"[^A-Z0-9]", "", text)

    def _read_csv(self, input_path: Path, **kwargs) -> pd.DataFrame:
        last_error: Optional[Exception] = None
        for encoding in ("utf-8-sig", "latin-1"):
            try:
                if "on_bad_lines" not in kwargs:
                    return pd.read_csv(
                        input_path,
                        sep=None,
                        engine="python",
                        encoding=encoding,
                        on_bad_lines="error",
                        **kwargs,
                    )
                return pd.read_csv(input_path, sep=None, engine="python", encoding=encoding, **kwargs)
            except pd.errors.ParserError as exc:
                try:
                    logger.warning("CSV malformato, salto righe non parseabili: %s", exc)
                    return pd.read_csv(
                        input_path,
                        sep=None,
                        engine="python",
                        encoding=encoding,
                        on_bad_lines="skip",
                        **kwargs,
                    )
                except Exception as exc2:
                    last_error = exc2
            except Exception as exc:
                last_error = exc
        if last_error:
            raise last_error
        raise ValueError("Impossibile leggere il file CSV")

    def _read_tabular(self, input_path: Path, **kwargs) -> pd.DataFrame:
        if input_path.suffix.lower() == ".csv":
            return self._read_csv(input_path, **kwargs)
        return pd.read_excel(input_path, **kwargs)

    
    def _parse_date_series(self, series: pd.Series) -> pd.Series:
        if pd.api.types.is_datetime64_any_dtype(series):
            return series
        as_text = series.astype(str).str.strip()
        iso_mask = as_text.str.match(r"^\d{4}-\d{2}-\d{2}$")
        parsed_iso = pd.to_datetime(series.where(iso_mask), errors="coerce", dayfirst=False)
        parsed_other = pd.to_datetime(series.where(~iso_mask), errors="coerce", dayfirst=True)
        return parsed_iso.fillna(parsed_other)

    def _extract_quarter_period(self, input_path: Path) -> Optional[tuple[pd.Timestamp, pd.Timestamp]]:
        name = input_path.name.upper()
        match = re.search(r"(I{1,3}|IV|[1-4])\s*°?\s*TRIM(?:ESTRE)?\.?\s*[-_ ]*(20\d{2})", name)
        if not match:
            return None
        quarter_raw, year_raw = match.groups()
        quarter_map = {"I": 1, "II": 2, "III": 3, "IV": 4}
        quarter = quarter_map[quarter_raw] if quarter_raw in quarter_map else int(quarter_raw)
        year = int(year_raw)
        month_start = (quarter - 1) * 3 + 1
        start = pd.Timestamp(year=year, month=month_start, day=1)
        end = start + pd.offsets.QuarterEnd()
        return start, end

    def _extract_quarter_period_extended_for_pagato(self, input_path: Path) -> Optional[tuple[pd.Timestamp, pd.Timestamp]]:
        period = self._extract_quarter_period(input_path)
        if period is None:
            return None
        start, end = period
        return start, end + pd.Timedelta(days=8)

    def _filter_by_file_quarter(self, df: pd.DataFrame, date_column: str, input_path: Path) -> pd.DataFrame:
        if date_column not in df.columns:
            return df
        period = self._extract_quarter_period(input_path)
        if period is None:
            return df
        start, end = period
        date_series = self._parse_date_series(df[date_column])
        return df[date_series.between(start, end)].copy()

    def _filter_by_file_quarter_extended(self, df: pd.DataFrame, date_column: str, input_path: Path) -> pd.DataFrame:
        if date_column not in df.columns:
            return df
        period = self._extract_quarter_period_extended_for_pagato(input_path)
        if period is None:
            return df
        start, end = period
        date_series = self._parse_date_series(df[date_column])
        return df[date_series.between(start, end)].copy()

    def _to_number_series(self, series: pd.Series) -> pd.Series:
        normalized = series.astype(str).str.replace(" ", "", regex=False).str.strip()
        has_dot = normalized.str.contains(r"\.", regex=True, na=False)
        has_comma = normalized.str.contains(",", regex=False, na=False)
        both_mask = has_dot & has_comma
        normalized = normalized.where(~both_mask, normalized.str.replace(".", "", regex=False))
        normalized = normalized.str.replace(",", ".", regex=False)
        return pd.to_numeric(normalized, errors="coerce")

    def _to_number_series_it(self, series: pd.Series) -> pd.Series:
        normalized = series.astype(str).str.replace(" ", "", regex=False).str.strip()
        normalized = normalized.str.replace(".", "", regex=False).str.replace(",", ".", regex=False)
        return pd.to_numeric(normalized, errors="coerce")

    def _load_nfs_compare_df(self, nfs_input_path: Path) -> pd.DataFrame:
        df = self._read_tabular(nfs_input_path, dtype=str)
        df.columns = [str(c).strip() for c in df.columns]

        rename_map: dict[str, str] = {}
        for col in df.columns:
            norm = self._normalize_col_name(col)
            mapped = {
                "RAGIONESOCIALE": "C_NOME",
                "FTPROT": "FAT_PROT",
                "FTPROTOCOLLO": "FAT_PROT",
                "FT_PROT": "FAT_PROT",
                "FATPROT": "FAT_PROT",
                "NPROTOCOLLO": "FAT_NUM",
                "NUMEROPROTOCOLLO": "FAT_NUM",
                "PROTMAND": "FAT_NUM",
                "PROT_MAND": "FAT_NUM",
                "NFATTURE": "FAT_NDOC",
                "NDOCUMENTO": "FAT_NDOC",
                "N_DOCUMENTO": "FAT_NDOC",
                "DATAFATTURA": "FAT_DATDOC",
                "DATAFATTURE": "FAT_DATDOC",
                "DATA_FATTURA": "FAT_DATDOC",
                "DATAREGISTRAZIONE": "FAT_DATREG",
                "DATAREGFATTURA": "FAT_DATREG",
                "DATA_REG_FATTURA": "FAT_DATREG",
                "FAT_DATREG": "FAT_DATREG",
                "IMPOSTA": "FAT_TOTIVA",
                "IMPTOTIVA": "FAT_TOTIVA",
                "IMP_TOT_IVA": "FAT_TOTIVA",
                "TOTIMPONIBILE": "IMPONIBILE",
                "IMPONIBILE": "IMPONIBILE",
                "IMP_TOT_FATTURA": "FAT_TOTFAT",
                "RITCODICETRIBUTO": "RA_CODTRIB",
                "CODTRIBUTO": "RA_CODTRIB",
                "COD_TRIBUTO": "RA_CODTRIB",
                "RA_CODTRIB": "RA_CODTRIB",
                "IDENTIFICATIVOSDI": "TMC_G8",
                "IDENTSDI": "TMC_G8",
                "TMCG8": "TMC_G8",
                "TMC_G8": "TMC_G8",
                "IDENT_SDI": "TMC_G8",
                "VALUTADELMANDATO": "M2_TMC_DATREG",
                "M2TMCDATREG": "M2_TMC_DATREG",
                "M2_TMC_DATREG": "M2_TMC_DATREG",
                "FAT_DATDOC": "FAT_DATDOC",
                "C_NOME": "C_NOME",
                "FAT_NDOC": "FAT_NDOC",
                "FAT_PROT": "FAT_PROT",
                "FAT_NUM": "FAT_NUM",
            }.get(norm)
            if mapped and mapped != col:
                rename_map[col] = mapped

        if rename_map:
            df = df.rename(columns=rename_map)

        if "FAT_NUM" not in df.columns and "FAT_NDOC" in df.columns:
            df["FAT_NUM"] = df["FAT_NDOC"]
        if "FAT_NDOC" not in df.columns and "FAT_NUM" in df.columns:
            df["FAT_NDOC"] = df["FAT_NUM"]

        for col, default in self.NFS_OPTIONAL_DEFAULTS.items():
            if col not in df.columns:
                df[col] = default

        missing = [c for c in self.NFS_REQUIRED_COLUMNS if c not in df.columns]
        if missing:
            raise ValueError(f"Colonne mancanti nel file NFS: {', '.join(missing)}")

        return df

    def _load_pisa_compare_df(self, pisa_input_path: Path) -> pd.DataFrame:
        df_pisa_raw = self._read_tabular(pisa_input_path, dtype=str)
        rename_map: dict[str, str] = {}

        if "Numero fattura" not in df_pisa_raw.columns and "C" in df_pisa_raw.columns:
            rename_map["C"] = "Numero fattura"
        if "Data emissione" not in df_pisa_raw.columns:
            if "F" in df_pisa_raw.columns:
                rename_map["F"] = "Data emissione"
            elif "Data pagamento" in df_pisa_raw.columns:
                rename_map["Data pagamento"] = "Data emissione"
        if "Importo fattura" not in df_pisa_raw.columns:
            if "Importo Fattura" in df_pisa_raw.columns:
                rename_map["Importo Fattura"] = "Importo fattura"
            elif "Importo" in df_pisa_raw.columns:
                rename_map["Importo"] = "Importo fattura"

        if rename_map:
            df_pisa_raw = df_pisa_raw.rename(columns=rename_map)

        if "Importo pagato" in df_pisa_raw.columns:
            df_pisa_raw["Importo fattura"] = df_pisa_raw["Importo pagato"]

        missing_pisa = [col for col in self.PISA_REQUIRED_COLUMNS if col not in df_pisa_raw.columns]
        if missing_pisa:
            raise ValueError(f"Colonne mancanti nel file Pisa: {', '.join(missing_pisa)}")
        return df_pisa_raw[self.PISA_REQUIRED_COLUMNS].copy()

    def process_files(self, nfs_input_path: Path, pisa_input_path: Path, output_path: Path) -> Dict[str, Any]:
        df_nfs_raw = self._load_nfs_compare_df(nfs_input_path)
        df_pisa = self._load_pisa_compare_df(pisa_input_path)

        nfs_period = self._extract_quarter_period(nfs_input_path)
        pisa_period = self._extract_quarter_period(pisa_input_path)
        if nfs_period and pisa_period and nfs_period != pisa_period:
            raise ValueError("I due file appartengono a trimestri/anni diversi in base al nome file.")

        df_nfs_lookup = df_nfs_raw[["FAT_DATREG", "TMC_G8"]].copy()
        df_nfs_lookup.rename(columns={"FAT_DATREG": "Datat reg.", "TMC_G8": "Identificativo SDI"}, inplace=True)
        df_nfs_lookup["Datat reg."] = self._parse_date_series(df_nfs_lookup["Datat reg."])
        df_nfs_lookup["_SDI_KEY"] = self._normalize_sdi(df_nfs_lookup["Identificativo SDI"])

        # Nuova procedura: teniamo solo protocolli ammessi e deduplica su 3 campi.
        nfs_protocol_raw = df_nfs_raw["FAT_PROT"].astype(str).str.strip().str.upper()
        df_nfs_filtered = df_nfs_raw[nfs_protocol_raw.isin(self.NFS_ALLOWED_PROTOCOLS)].copy()
        df_nfs_deduped = df_nfs_filtered.drop_duplicates(subset=["FAT_NDOC", "FAT_DATDOC", "C_NOME"]).copy()
        df_nfs = df_nfs_deduped[self.NFS_REQUIRED_COLUMNS].copy()
        df_nfs.rename(columns=self.NFS_RENAME_MAP, inplace=True)
        df_nfs["Data Fatture"] = self._parse_date_series(df_nfs["Data Fatture"])
        df_nfs["Datat reg."] = self._parse_date_series(df_nfs["Datat reg."])
        df_nfs["Imponibile"] = self._to_number_series_it(df_nfs["Imponibile"]).fillna(0)
        df_nfs["Importo Pagamento"] = df_nfs["Imponibile"]

        df_pisa["Data emissione"] = self._parse_date_series(df_pisa["Data emissione"])
        df_pisa["Importo fattura"] = self._to_number_series(df_pisa["Importo fattura"]).fillna(0)

        df_nfs["_SDI_KEY"] = self._normalize_sdi(df_nfs["Identificativo SDI"])
        df_pisa["_SDI_KEY"] = self._normalize_sdi(df_pisa["Identificativo SDI"])

        nfs_protocol_series = df_nfs["Prot."].astype(str).str.strip().str.upper()
        nfs_cart_mask = nfs_protocol_series.isin(self.NFS_CARTACEE_PROTOCOLS)
        nfs_elet_mask = nfs_protocol_series.isin(self.NFS_ELETTRONICHE_PROTOCOLS | self.NFS_AUTOFATTURE_PROTOCOLS)
        pisa_cart_mask = self._is_empty_sdi(df_pisa["_SDI_KEY"])

        nfs_cart_count = int(nfs_cart_mask.sum())
        nfs_elet_count = int(nfs_elet_mask.sum())
        pisa_cart_count = int(pisa_cart_mask.sum())
        pisa_elet_count = int((~pisa_cart_mask).sum())

        nfs_cart_amount = round(float(df_nfs.loc[nfs_cart_mask, "Importo Pagamento"].sum()), 2)
        nfs_elet_amount = round(float(df_nfs.loc[nfs_elet_mask, "Importo Pagamento"].sum()), 2)
        pisa_cart_amount = round(float(df_pisa.loc[pisa_cart_mask, "Importo fattura"].sum()), 2)
        pisa_elet_amount = round(float(df_pisa.loc[~pisa_cart_mask, "Importo fattura"].sum()), 2)

        period_label = "Tutto il periodo"
        active_period = nfs_period or pisa_period
        if active_period:
            start, end = active_period
            period_label = f"{start.strftime('%d/%m/%Y')} - {end.strftime('%d/%m/%Y')}"

        summary = {
            "period": period_label,
            "nfs": {
                "cartacee": {"count": nfs_cart_count, "amount": nfs_cart_amount, "amount_column": "IMPONIBILE"},
                "elettroniche": {"count": nfs_elet_count, "amount": nfs_elet_amount, "amount_column": "IMPONIBILE"},
            },
            "pisa": {
                "cartacee": {"count": pisa_cart_count, "amount": pisa_cart_amount, "amount_column": "Importo pagato"},
                "elettroniche": {"count": pisa_elet_count, "amount": pisa_elet_amount, "amount_column": "Importo pagato"},
            },
        }
        for side in ("nfs", "pisa"):
            for kind in ("cartacee", "elettroniche"):
                summary[side][kind]["imponibile"] = summary[side][kind]["amount"]

        wb = Workbook()
        wb.remove(wb.active)

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        self._create_confronto_sheet(
            wb=wb,
            nfs_cart_count=nfs_cart_count,
            nfs_cart_amount=nfs_cart_amount,
            nfs_elet_count=nfs_elet_count,
            nfs_elet_amount=nfs_elet_amount,
            pisa_cart_count=pisa_cart_count,
            pisa_cart_amount=pisa_cart_amount,
            pisa_elet_count=pisa_elet_count,
            pisa_elet_amount=pisa_elet_amount,
            header_fill=header_fill,
            header_font=header_font,
        )
        self._create_fatture_da_verificare_sheet(
            wb=wb,
            df_nfs=df_nfs,
            df_pisa=df_pisa,
            header_fill=header_fill,
            header_font=header_font,
        )

        wb.save(output_path)
        return summary

    def _is_empty_sdi(self, series: pd.Series) -> pd.Series:
        # Regola metodologica: vuoto = cartacea, pieno = elettronica.
        normalized = series.astype(str).str.strip().where(~series.isna(), "")
        return normalized.eq("")

    def _create_confronto_sheet(
        self,
        wb: Workbook,
        nfs_cart_count: int,
        nfs_cart_amount: float,
        nfs_elet_count: int,
        nfs_elet_amount: float,
        pisa_cart_count: int,
        pisa_cart_amount: float,
        pisa_elet_count: int,
        pisa_elet_amount: float,
        header_fill: PatternFill,
        header_font: Font,
    ) -> None:
        ws = wb.create_sheet("Confronto")
        total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        total_font = Font(bold=True)

        headers = [
            "Categoria",
            "NFS Numero",
            "NFS Importo",
            "Pisa Numero",
            "Pisa Importo",
            "Delta Numero",
            "Delta Importo",
        ]
        for col_idx, value in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=value)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        rows = [
            ("Cartacee", nfs_cart_count, nfs_cart_amount, pisa_cart_count, pisa_cart_amount),
            ("Elettroniche", nfs_elet_count, nfs_elet_amount, pisa_elet_count, pisa_elet_amount),
            (
                "Totale",
                nfs_cart_count + nfs_elet_count,
                round(nfs_cart_amount + nfs_elet_amount, 2),
                pisa_cart_count + pisa_elet_count,
                round(pisa_cart_amount + pisa_elet_amount, 2),
            ),
        ]
        money_format = "#,##0.00"
        for row_idx, (categoria, n_num, n_imp, p_num, p_imp) in enumerate(rows, start=2):
            ws.cell(row=row_idx, column=1, value=categoria)
            ws.cell(row=row_idx, column=2, value=n_num)
            ws.cell(row=row_idx, column=3, value=n_imp).number_format = money_format
            ws.cell(row=row_idx, column=4, value=p_num)
            ws.cell(row=row_idx, column=5, value=p_imp).number_format = money_format
            ws.cell(row=row_idx, column=6, value=n_num - p_num)
            ws.cell(row=row_idx, column=7, value=round(n_imp - p_imp, 2)).number_format = money_format

        for cell in ws[ws.max_row]:
            cell.fill = total_fill
            cell.font = total_font

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 14
        ws.column_dimensions["G"].width = 16

    def _normalize_sdi(self, series: pd.Series) -> pd.Series:
        def normalize_value(value: Any) -> str:
            if pd.isna(value):
                return ""
            if isinstance(value, int):
                return str(value)
            if isinstance(value, float):
                if value.is_integer():
                    return str(int(value))
                return str(value).strip()
            text = str(value).strip()
            match = re.fullmatch(r"(\d+)\.0+", text)
            if match:
                return match.group(1)
            return text

        return series.map(normalize_value)

    def _create_fatture_da_verificare_sheet(
        self,
        wb: Workbook,
        df_nfs: pd.DataFrame,
        df_pisa: pd.DataFrame,
        header_fill: PatternFill,
        header_font: Font,
    ) -> None:
        ws = wb.create_sheet("Differenze tra file")

        headers = [
            "Esito",
            "Identificativo SDI",
            "NFS Ragione sociale",
            "NFS N.fatture",
            "NFS Datat reg.",
            "NFS Prot.",
            "NFS Importo",
            "Pisa Creditore",
            "Pisa Numero fattura",
            "Pisa Data emissione",
            "Pisa Importo fattura",
            "Delta Importo",
        ]
        for col_idx, value in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=value)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        nfs_sdi_empty = self._is_empty_sdi(df_nfs["_SDI_KEY"])
        pisa_sdi_empty = self._is_empty_sdi(df_pisa["_SDI_KEY"])

        # Differenze elettroniche: confrontiamo SOLO le elettroniche NFS (da protocollo)
        # contro le elettroniche Pisa (SDI pieno), in modo coerente col foglio "Confronto".
        nfs_protocol_series = df_nfs["Prot."].astype(str).str.strip().str.upper()
        nfs_elet_mask = nfs_protocol_series.isin(self.NFS_ELETTRONICHE_PROTOCOLS | self.NFS_AUTOFATTURE_PROTOCOLS)
        nfs_elet = df_nfs[nfs_elet_mask].copy()
        nfs_elet_sdi_empty = self._is_empty_sdi(nfs_elet["_SDI_KEY"])
        nfs_non_empty = nfs_elet[~nfs_elet_sdi_empty].copy()

        pisa_non_empty = df_pisa[~pisa_sdi_empty].copy()

        money_format = "#,##0.00"
        date_format = "dd/mm/yyyy"

        nfs_non_empty["_SDI_KEY"] = nfs_non_empty["_SDI_KEY"].astype(str).str.strip()
        pisa_non_empty["_SDI_KEY"] = pisa_non_empty["_SDI_KEY"].astype(str).str.strip()

        from collections import Counter, defaultdict

        nfs_sdi_list = nfs_non_empty["_SDI_KEY"].astype(str).str.strip().tolist()
        pisa_sdi_list = pisa_non_empty["_SDI_KEY"].astype(str).str.strip().tolist()

        nfs_sdi_counter = Counter([v for v in nfs_sdi_list if v != ""])
        pisa_sdi_counter = Counter([v for v in pisa_sdi_list if v != ""])

        extra_pisa_sdi = list((pisa_sdi_counter - nfs_sdi_counter).elements())
        extra_nfs_sdi = list((nfs_sdi_counter - pisa_sdi_counter).elements())

        pisa_by_sdi: dict[str, list[pd.Series]] = defaultdict(list)
        for _, r in pisa_non_empty.sort_values(by=["Data emissione", "Numero fattura"], na_position="last").iterrows():
            key = str(r.get("_SDI_KEY", "")).strip()
            if key:
                pisa_by_sdi[key].append(r)

        nfs_by_sdi: dict[str, list[pd.Series]] = defaultdict(list)
        for _, r in nfs_non_empty.sort_values(by=["Datat reg.", "N.fatture"], na_position="last").iterrows():
            key = str(r.get("_SDI_KEY", "")).strip()
            if key:
                nfs_by_sdi[key].append(r)

        row_idx = 2

        pisa_elet_total = int((~pisa_sdi_empty).sum())
        nfs_elet_total = int(nfs_elet_mask.sum())
        delta_elet = int(pisa_elet_total - nfs_elet_total)
        if delta_elet > 0:
            to_show_pisa = extra_pisa_sdi[:delta_elet]
            to_show_nfs: list[str] = []
        elif delta_elet < 0:
            to_show_pisa = []
            to_show_nfs = extra_nfs_sdi[: abs(delta_elet)]
        else:
            to_show_pisa = []
            to_show_nfs = []

        for sdi in to_show_pisa:
            pisa_row = pisa_by_sdi.get(sdi, [])
            pisa_item = pisa_row.pop(0) if pisa_row else None
            ws.cell(row=row_idx, column=1, value="Solo Pisa")
            ws.cell(row=row_idx, column=2, value=sdi)
            for c in range(3, 8):
                ws.cell(row=row_idx, column=c, value="")
            ws.cell(row=row_idx, column=8, value="" if pisa_item is None else pisa_item.get("Creditore", ""))
            ws.cell(row=row_idx, column=9, value="" if pisa_item is None else pisa_item.get("Numero fattura", ""))
            c10 = ws.cell(row=row_idx, column=10, value=None if pisa_item is None else pisa_item.get("Data emissione", None))
            if c10.value is not None:
                c10.number_format = date_format
            c11 = ws.cell(row=row_idx, column=11, value=0.0 if pisa_item is None else float(pisa_item.get("Importo fattura", 0.0)))
            c11.number_format = money_format
            c12 = ws.cell(row=row_idx, column=12, value=-float(pisa_item.get("Importo fattura", 0.0)) if pisa_item is not None else 0.0)
            c12.number_format = money_format
            row_idx += 1

        for sdi in to_show_nfs:
            nfs_row = nfs_by_sdi.get(sdi, [])
            nfs_item = nfs_row.pop(0) if nfs_row else None
            ws.cell(row=row_idx, column=1, value="Solo NFS")
            ws.cell(row=row_idx, column=2, value=sdi)
            ws.cell(row=row_idx, column=3, value="" if nfs_item is None else nfs_item.get("Ragione sociale", ""))
            ws.cell(row=row_idx, column=4, value="" if nfs_item is None else nfs_item.get("N.fatture", ""))
            c5 = ws.cell(row=row_idx, column=5, value=None if nfs_item is None else nfs_item.get("Datat reg.", None))
            if c5.value is not None:
                c5.number_format = date_format
            ws.cell(row=row_idx, column=6, value="" if nfs_item is None else nfs_item.get("Prot.", ""))
            c7 = ws.cell(row=row_idx, column=7, value=0.0 if nfs_item is None else float(nfs_item.get("Importo Pagamento", 0.0)))
            c7.number_format = money_format
            for c in range(8, 12):
                ws.cell(row=row_idx, column=c, value="")
            c12 = ws.cell(row=row_idx, column=12, value=float(nfs_item.get("Importo Pagamento", 0.0)) if nfs_item is not None else 0.0)
            c12.number_format = money_format
            row_idx += 1

        # Cartacee: nel riepilogo "Confronto" la differenza cartacee deriva dalla classificazione NFS per protocollo
        # e Pisa per SDI vuoto. Qui mostriamo SOLO le fatture cartacee che generano differenza (non tutte).
        nfs_cart = df_nfs[df_nfs["Prot."].astype(str).str.strip().str.upper().isin(self.NFS_CARTACEE_PROTOCOLS)].copy()
        pisa_cart = df_pisa[pisa_sdi_empty].copy()

        def normalize_text(value: Any) -> str:
            if pd.isna(value):
                return ""
            return str(value).strip().upper()

        def normalize_date(value: Any) -> str:
            if value is None or (isinstance(value, float) and pd.isna(value)):
                return ""
            try:
                dt = pd.to_datetime(value, errors="coerce", dayfirst=True)
            except Exception:
                dt = pd.NaT
            if pd.isna(dt):
                return ""
            return dt.strftime("%Y-%m-%d")

        def normalize_amount(value: Any) -> str:
            try:
                num = float(pd.to_numeric(pd.Series([value]), errors="coerce").fillna(0).iloc[0])
            except Exception:
                num = 0.0
            return f"{num:.2f}"

        def make_cart_key(source: str, row: pd.Series) -> str:
            if source == "nfs":
                num = normalize_text(row.get("N.fatture", ""))
                date = normalize_date(row.get("Data Fatture", None))
                amt = normalize_amount(row.get("Importo Pagamento", 0.0))
                name = normalize_text(row.get("Ragione sociale", ""))
            else:
                num = normalize_text(row.get("Numero fattura", ""))
                date = normalize_date(row.get("Data emissione", None))
                amt = normalize_amount(row.get("Importo fattura", 0.0))
                name = normalize_text(row.get("Creditore", ""))
            return f"{num}|{date}|{amt}|{name}"

        from collections import Counter

        nfs_cart_keys = [make_cart_key("nfs", r) for _, r in nfs_cart.iterrows()]
        pisa_cart_keys = [make_cart_key("pisa", r) for _, r in pisa_cart.iterrows()]
        nfs_counter = Counter(nfs_cart_keys)
        pisa_counter = Counter(pisa_cart_keys)

        only_nfs_keys = list((nfs_counter - pisa_counter).elements())
        only_pisa_keys = list((pisa_counter - nfs_counter).elements())

        delta_cart = int(len(nfs_cart) - int(pisa_sdi_empty.sum()))
        if delta_cart > 0:
            to_show_nfs_keys = only_nfs_keys[:delta_cart]
            to_show_pisa_keys: list[str] = []
        elif delta_cart < 0:
            to_show_nfs_keys = []
            to_show_pisa_keys = only_pisa_keys[: abs(delta_cart)]
        else:
            to_show_nfs_keys = []
            to_show_pisa_keys = []

        if to_show_nfs_keys:
            key_set = set(to_show_nfs_keys)
            for _, nfs_row in nfs_cart.iterrows():
                k = make_cart_key("nfs", nfs_row)
                if k not in key_set:
                    continue
                key_set.remove(k)
                ws.cell(row=row_idx, column=1, value="Solo NFS (Cartacea)")
                ws.cell(row=row_idx, column=2, value="")
                ws.cell(row=row_idx, column=3, value=nfs_row.get("Ragione sociale", ""))
                ws.cell(row=row_idx, column=4, value=nfs_row.get("N.fatture", ""))
                c5 = ws.cell(row=row_idx, column=5, value=nfs_row.get("Datat reg.", None))
                if c5.value is not None:
                    c5.number_format = date_format
                ws.cell(row=row_idx, column=6, value=nfs_row.get("Prot.", ""))
                c7 = ws.cell(row=row_idx, column=7, value=float(nfs_row.get("Importo Pagamento", 0.0)))
                c7.number_format = money_format
                for c in range(8, 12):
                    ws.cell(row=row_idx, column=c, value="")
                c12 = ws.cell(row=row_idx, column=12, value=float(nfs_row.get("Importo Pagamento", 0.0)))
                c12.number_format = money_format
                row_idx += 1

        if to_show_pisa_keys:
            key_set = set(to_show_pisa_keys)
            for _, pisa_row in pisa_cart.iterrows():
                k = make_cart_key("pisa", pisa_row)
                if k not in key_set:
                    continue
                key_set.remove(k)
                ws.cell(row=row_idx, column=1, value="Solo Pisa (Cartacea)")
                ws.cell(row=row_idx, column=2, value="")
                for c in range(3, 8):
                    ws.cell(row=row_idx, column=c, value="")
                ws.cell(row=row_idx, column=8, value=pisa_row.get("Creditore", ""))
                ws.cell(row=row_idx, column=9, value=pisa_row.get("Numero fattura", ""))
                c10 = ws.cell(row=row_idx, column=10, value=pisa_row.get("Data emissione", None))
                if c10.value is not None:
                    c10.number_format = date_format
                c11 = ws.cell(row=row_idx, column=11, value=float(pisa_row.get("Importo fattura", 0.0)))
                c11.number_format = money_format
                c12 = ws.cell(row=row_idx, column=12, value=-float(pisa_row.get("Importo fattura", 0.0)))
                c12.number_format = money_format
                row_idx += 1

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 26
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 16
        ws.column_dimensions["H"].width = 26
        ws.column_dimensions["I"].width = 16
        ws.column_dimensions["J"].width = 16
        ws.column_dimensions["K"].width = 18
        ws.column_dimensions["L"].width = 16

    def _create_differenze_elettroniche_sheet(
        self,
        wb: Workbook,
        df_nfs: pd.DataFrame,
        df_pisa: pd.DataFrame,
        nfs_elet_mask: pd.Series,
        pisa_cart_mask: pd.Series,
        header_fill: PatternFill,
        header_font: Font,
    ) -> None:
        ws = wb.create_sheet("Differenze Elettroniche SDI")

        headers = [
            "Sezione",
            "Identificativo SDI",
            "NFS Ragione sociale",
            "NFS N.fatture",
            "NFS Datat reg.",
            "NFS Prot.",
            "NFS Imponibile",
            "Pisa Creditore",
            "Pisa Numero fattura",
            "Pisa Data emissione",
            "Pisa Importo fattura",
        ]
        for col_idx, value in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=value)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        money_format = "#,##0.00"
        date_format = "dd/mm/yyyy"

        nfs_elet = df_nfs[nfs_elet_mask].copy()
        pisa_elet = df_pisa[~pisa_cart_mask].copy()

        nfs_sdi_empty = self._is_empty_sdi(nfs_elet["_SDI_KEY"])
        pisa_sdi_empty = self._is_empty_sdi(pisa_elet["_SDI_KEY"])
        nfs_elet_non_empty = nfs_elet[~nfs_sdi_empty].copy()
        pisa_elet_non_empty = pisa_elet[~pisa_sdi_empty].copy()

        nfs_keys = set(nfs_elet_non_empty["_SDI_KEY"].astype(str).str.strip())
        pisa_keys = set(pisa_elet_non_empty["_SDI_KEY"].astype(str).str.strip())

        only_pisa_keys = sorted(pisa_keys - nfs_keys)
        only_nfs_keys = sorted(nfs_keys - pisa_keys)

        nfs_elet_empty_sdi = nfs_elet[nfs_sdi_empty].copy()

        row_idx = 2

        def write_row(
            section: str,
            sdi: str,
            nfs_row: Optional[pd.Series],
            pisa_row: Optional[pd.Series],
        ) -> None:
            nonlocal row_idx
            ws.cell(row=row_idx, column=1, value=section)
            ws.cell(row=row_idx, column=2, value=sdi)

            if nfs_row is not None:
                ws.cell(row=row_idx, column=3, value=nfs_row.get("Ragione sociale", ""))
                ws.cell(row=row_idx, column=4, value=nfs_row.get("N.fatture", ""))
                c5 = ws.cell(row=row_idx, column=5, value=nfs_row.get("Datat reg.", None))
                if c5.value is not None:
                    c5.number_format = date_format
                ws.cell(row=row_idx, column=6, value=nfs_row.get("Prot.", ""))
                c7 = ws.cell(row=row_idx, column=7, value=float(nfs_row.get("Imponibile", 0.0)))
                c7.number_format = money_format
            else:
                for c in range(3, 8):
                    ws.cell(row=row_idx, column=c, value="")

            if pisa_row is not None:
                ws.cell(row=row_idx, column=8, value=pisa_row.get("Creditore", ""))
                ws.cell(row=row_idx, column=9, value=pisa_row.get("Numero fattura", ""))
                c10 = ws.cell(row=row_idx, column=10, value=pisa_row.get("Data emissione", None))
                if c10.value is not None:
                    c10.number_format = date_format
                c11 = ws.cell(row=row_idx, column=11, value=float(pisa_row.get("Importo fattura", 0.0)))
                c11.number_format = money_format
            else:
                for c in range(8, 12):
                    ws.cell(row=row_idx, column=c, value="")

            row_idx += 1

        nfs_first_by_key = (
            nfs_elet_non_empty.sort_values(by=["Datat reg.", "N.fatture"], na_position="last")
            .drop_duplicates(subset=["_SDI_KEY"], keep="first")
            .set_index("_SDI_KEY")
        )
        pisa_first_by_key = (
            pisa_elet_non_empty.sort_values(by=["Data emissione", "Numero fattura"], na_position="last")
            .drop_duplicates(subset=["_SDI_KEY"], keep="first")
            .set_index("_SDI_KEY")
        )

        for key in only_pisa_keys:
            write_row("Solo Pisa", key, None, pisa_first_by_key.loc[key])

        for key in only_nfs_keys:
            write_row("Solo NFS", key, nfs_first_by_key.loc[key], None)

        for _, r in nfs_elet_empty_sdi.sort_values(by=["Datat reg.", "N.fatture"], na_position="last").iterrows():
            write_row("NFS SDI vuoto", "", r, None)

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 26
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 16
        ws.column_dimensions["H"].width = 26
        ws.column_dimensions["I"].width = 16
        ws.column_dimensions["J"].width = 16
        ws.column_dimensions["K"].width = 18

    def _create_differenze_sdi_univoche_sheet(
        self,
        wb: Workbook,
        df_nfs: pd.DataFrame,
        df_pisa: pd.DataFrame,
        nfs_elet_mask: pd.Series,
        pisa_cart_mask: pd.Series,
        header_fill: PatternFill,
        header_font: Font,
    ) -> None:
        ws = wb.create_sheet("Differenze SDI in Comune")

        headers = [
            "Identificativo SDI",
            "NFS Ragione sociale",
            "NFS N.fatture",
            "NFS Datat reg.",
            "NFS Imponibile",
            "Pisa Creditore",
            "Pisa Numero fattura",
            "Pisa Data emissione",
            "Pisa Importo fattura",
            "Delta Importo",
        ]
        for col_idx, value in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=value)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        money_format = "#,##0.00"
        date_format = "dd/mm/yyyy"

        nfs_elet = df_nfs[nfs_elet_mask].copy()
        pisa_elet = df_pisa[~pisa_cart_mask].copy()

        nfs_sdi_empty = self._is_empty_sdi(nfs_elet["_SDI_KEY"])
        pisa_sdi_empty = self._is_empty_sdi(pisa_elet["_SDI_KEY"])
        nfs_elet = nfs_elet[~nfs_sdi_empty].copy()
        pisa_elet = pisa_elet[~pisa_sdi_empty].copy()

        nfs_counts = nfs_elet["_SDI_KEY"].value_counts()
        pisa_counts = pisa_elet["_SDI_KEY"].value_counts()

        common_keys = set(nfs_counts.index) & set(pisa_counts.index)
        common_unique_keys = sorted(
            [k for k in common_keys if int(nfs_counts.get(k, 0)) == 1 and int(pisa_counts.get(k, 0)) == 1]
        )

        nfs_unique = nfs_elet.set_index("_SDI_KEY", drop=False)
        pisa_unique = pisa_elet.set_index("_SDI_KEY", drop=False)

        row_idx = 2
        for key in common_unique_keys:
            nfs_row = nfs_unique.loc[key]
            pisa_row = pisa_unique.loc[key]

            delta = round(float(nfs_row.get("Imponibile", 0.0)) - float(pisa_row.get("Importo fattura", 0.0)), 2)
            if abs(delta) <= 0.01:
                continue

            ws.cell(row=row_idx, column=1, value=key)
            ws.cell(row=row_idx, column=2, value=nfs_row.get("Ragione sociale", ""))
            ws.cell(row=row_idx, column=3, value=nfs_row.get("N.fatture", ""))
            c4 = ws.cell(row=row_idx, column=4, value=nfs_row.get("Datat reg.", None))
            if c4.value is not None:
                c4.number_format = date_format
            c5 = ws.cell(row=row_idx, column=5, value=float(nfs_row.get("Imponibile", 0.0)))
            c5.number_format = money_format

            ws.cell(row=row_idx, column=6, value=pisa_row.get("Creditore", ""))
            ws.cell(row=row_idx, column=7, value=pisa_row.get("Numero fattura", ""))
            c8 = ws.cell(row=row_idx, column=8, value=pisa_row.get("Data emissione", None))
            if c8.value is not None:
                c8.number_format = date_format
            c9 = ws.cell(row=row_idx, column=9, value=float(pisa_row.get("Importo fattura", 0.0)))
            c9.number_format = money_format

            c10 = ws.cell(row=row_idx, column=10, value=delta)
            c10.number_format = money_format
            row_idx += 1

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 26
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 26
        ws.column_dimensions["G"].width = 16
        ws.column_dimensions["H"].width = 16
        ws.column_dimensions["I"].width = 18
        ws.column_dimensions["J"].width = 16

    def _create_pisa_solo_mese_nfs_sheet(
        self,
        wb: Workbook,
        df_nfs_lookup: pd.DataFrame,
        df_nfs_jan: pd.DataFrame,
        df_pisa_jan: pd.DataFrame,
        nfs_elet_mask: pd.Series,
        pisa_cart_mask: pd.Series,
        header_fill: PatternFill,
        header_font: Font,
    ) -> None:
        ws = wb.create_sheet("Pisa Solo - Mese NFS")

        headers = [
            "Identificativo SDI",
            "Pisa Creditore",
            "Pisa Numero fattura",
            "Pisa Data emissione",
            "Pisa Importo fattura",
            "NFS Mesi trovati",
            "NFS Prima registrazione",
        ]
        for col_idx, value in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=value)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        money_format = "#,##0.00"
        date_format = "dd/mm/yyyy"

        pisa_elet = df_pisa_jan[~pisa_cart_mask].copy()
        pisa_elet = pisa_elet[~self._is_empty_sdi(pisa_elet["_SDI_KEY"])].copy()
        nfs_elet = df_nfs_jan[nfs_elet_mask].copy()
        nfs_elet = nfs_elet[~self._is_empty_sdi(nfs_elet["_SDI_KEY"])].copy()

        pisa_keys = set(pisa_elet["_SDI_KEY"].astype(str).str.strip())
        nfs_keys = set(nfs_elet["_SDI_KEY"].astype(str).str.strip())
        only_pisa_keys = sorted(pisa_keys - nfs_keys)

        pisa_first_by_key = (
            pisa_elet.sort_values(by=["Data emissione", "Numero fattura"], na_position="last")
            .drop_duplicates(subset=["_SDI_KEY"], keep="first")
            .set_index("_SDI_KEY")
        )

        df_nfs_lookup_non_empty = df_nfs_lookup[~self._is_empty_sdi(df_nfs_lookup["_SDI_KEY"])].copy()
        df_nfs_lookup_non_empty["_SDI_KEY_NORM"] = df_nfs_lookup_non_empty["_SDI_KEY"].astype(str).str.strip()
        df_nfs_lookup_non_empty["_NFS_DATE"] = pd.to_datetime(
            df_nfs_lookup_non_empty["Datat reg."], errors="coerce", dayfirst=True
        )
        df_nfs_lookup_non_empty["_NFS_MONTH"] = df_nfs_lookup_non_empty["_NFS_DATE"].dt.to_period("M").astype(str)
        nfs_months_by_key = (
            df_nfs_lookup_non_empty.dropna(subset=["_NFS_MONTH"])
            .groupby("_SDI_KEY_NORM")["_NFS_MONTH"]
            .agg(lambda values: sorted(set(values)))
        )
        nfs_first_reg_by_key = df_nfs_lookup_non_empty.groupby("_SDI_KEY_NORM")["_NFS_DATE"].min()

        row_idx = 2
        for key in only_pisa_keys:
            pisa_row = pisa_first_by_key.loc[key]
            months = nfs_months_by_key.get(key, [])
            first_reg = nfs_first_reg_by_key.get(key, None)

            ws.cell(row=row_idx, column=1, value=key)
            ws.cell(row=row_idx, column=2, value=pisa_row.get("Creditore", ""))
            ws.cell(row=row_idx, column=3, value=pisa_row.get("Numero fattura", ""))

            c4 = ws.cell(row=row_idx, column=4, value=pisa_row.get("Data emissione", None))
            if c4.value is not None:
                c4.number_format = date_format
            c5 = ws.cell(row=row_idx, column=5, value=float(pisa_row.get("Importo fattura", 0.0)))
            c5.number_format = money_format

            ws.cell(row=row_idx, column=6, value=", ".join(months))
            c7 = ws.cell(row=row_idx, column=7, value=first_reg)
            if c7.value is not None:
                c7.number_format = date_format

            row_idx += 1

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 22
        ws.column_dimensions["G"].width = 20
